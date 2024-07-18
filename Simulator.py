import time
from datetime import datetime
from math import *
from geopy.geocoders import Nominatim
import pandas as pd
from openpyxl import load_workbook, Workbook
from tkinter import *

def julian_day_centuary(year, month, day): 
    
    julianDay = (1461*(year + 4800 + (month/14)/12))/4 + (367*(month - 2 - 12 * ((month/14)/12)))/12 - (3*((year + 4900 + (month-14)/12)/100))/4 + day - 32075
    julianCentuary = (julianDay - 2451545)/36525
    return julianDay, julianCentuary

def get_time_past_local_midnight(hour, minutes , seconds):
    tplm = ((hour) + (minutes/60) + (seconds/3600))/24
    return tplm

def get_geomMeanLongAnom(julian_centuary):
    
    geo_mean_long = (280.46646 + julian_centuary*(36000.76983 + julian_centuary*0.0003032))%360
    geo_mean_anom = 357.52911 + julian_centuary*(35999.05029 - 0.0001537*julian_centuary)

    return geo_mean_long, geo_mean_anom

def get_escentEarthOrbit(julian_centuary):
    
    escenEarthOrbit = 0.016708634 - julian_centuary*(0.000042037 + 0.0000001267*julian_centuary)
    return escenEarthOrbit

def get_suntrueLong_suntrueAnom(geoAnom,geoMean ,julian_centuary):
    
    suneqctr = sin(radians(geoAnom))*(1.914602 - julian_centuary*(0.004817+0.000014*julian_centuary)) + sin(radians(2*geoAnom))*(0.019993-0.000101*julian_centuary) + sin(radians(3*geoAnom))*0.000289
    suntrueLong = geoMean + suneqctr
    suntrueAnom = geoAnom + suneqctr
    return suntrueLong, suntrueAnom

def get_sunRadVector_sunLongApp(escent, trueAnom, trueLong, juliancentuary):

    sunRadVector = (1.000001018*(1-escent*escent))/(1+escent*cos(radians(trueAnom)))
    sunLongApp = trueLong-0.00569-0.00478*sin(radians(125.04-1934.136*juliancentuary))
    return sunRadVector, sunLongApp

def get_mean_obliq_ecliptic_and_correction(julian_centuary):

    MOE = 23+(26+((21.448-julian_centuary*(46.815+julian_centuary*(0.00059-julian_centuary*0.001813))))/60)/60
    MOEcorr = MOE+0.00256*cos(radians(125.04-1934.136*julian_centuary))
    return MOE, MOEcorr

def get_sun_rt_Asccen_and_sun_decl(sunlongApp, MOEcorr):
    sunrtAsccen = degrees(atan2(cos(radians(sunlongApp)),cos(radians(MOEcorr))*sin(radians(sunlongApp))))
    sundeclination = degrees(asin(sin(radians(MOEcorr))*sin(radians(sunlongApp))))
    return sunrtAsccen, sundeclination

def get_var_y(MOEcorr):
    var_y = tan(radians(MOEcorr/2))*tan(radians(MOEcorr/2))
    return var_y

def get_EOTSN(var_y, geoMeanLong, geoMeanAnom, escent):
    
    EOTSN = 4*degrees(var_y*sin(2*radians(geoMeanLong))-2*escent*sin(radians(geoMeanAnom))+4*escent*var_y*sin(radians(geoMeanAnom))*cos(2*radians(geoMeanLong))-0.5*var_y*var_y*sin(4*radians(geoMeanLong))-1.25*escent*escent*sin(2*radians(geoMeanAnom)))
    return EOTSN

def HA_sunrise(sun_declination, latitude):
     
    ha_sunrise = degrees(acos(cos(radians(90.833))/(cos(radians(latitude))*cos(radians(sun_declination)))-tan(radians(latitude))*tan(radians(sun_declination))))
    return ha_sunrise

def solar_noon(EOTSN, longitude):
    
    timeoffset = 5.5
    solarNoon = (720-4*longitude-EOTSN+timeoffset*60)/1440
    solar_noon_hours = solarNoon*24
    solar_noon_minutes = ((solarNoon*24)-int((solarNoon*24)))*60
    solar_noon_timeString = str(int(solar_noon_hours)) + ":" + str(int(solar_noon_minutes))
    return solarNoon, solar_noon_timeString

def get_sunrise_sunset(solar_noon, ha_sunrise):
    
    sunrise = (solar_noon - ha_sunrise*4)/1440
    sunset = (solar_noon + ha_sunrise*4)/1440
    sunrise_hours = sunrise*24
    sunrise_minutes = ((sunrise*24) - int((sunrise*24)))*60
    sunrise_timeString = str(int(sunrise_hours)) + ":" + str(int(sunrise_minutes))

    sunset_hours = sunset*24
    sunset_minutes = ((sunset*24) - int((sunset*24)))*60
    sunset_timeString = str(int(sunset_hours)) + ":" + str(int(sunset_minutes))

    return sunrise_timeString, sunset_timeString, sunrise, sunset

def get_true_solar_time(tplm, EOT, longitude):
    
    true_solar_time = (tplm*1440 + EOT + 4*longitude - 60*5.5)%1440
    return true_solar_time

def get_EOT(d):
    B = 360*(d-81)/365
    EOT = 9.87*(sin(radians(2*B))) - 7.53*(cos(radians(B))) - 1.5*sin(radians(B))
    return EOT

def get_local_solar_time(timeInfo, EOT):
    LSTM = 15*(5.5)
    longitude = 72.54
    time_correction_factor = 240*(longitude - LSTM) + EOT
    local_solar_time = timeInfo[3]*3600 + timeInfo[4]*60 + timeInfo[5] + (time_correction_factor)
    return local_solar_time

def get_hourAngle(local_solar_time):
    hour_angle = (15/3600)*(local_solar_time - 12*3600)
    return hour_angle

def get_declinationAngle(d):
    declination_angle = -23.45*(cos(radians((360/365)*(d+10))))
    return declination_angle

def get_elevationAngle(declination, latitude, hourAngle):
    elevation_angle = degrees(asin((sin(radians(declination)) * sin(radians(latitude))) + (cos(radians(declination))*cos(radians(hourAngle))*cos(radians(latitude)))))
    return elevation_angle

def get_azimuthal(declination, latitude, HOURANGLE, elevationAngle):
    azimuthal = degrees(acos(((sin(radians(declination))*cos(radians(latitude))) - (cos(radians(declination))*sin(radians(latitude))*cos(radians(HOURANGLE))))/(cos(radians(elevationAngle)))))
    if HOURANGLE < 12:
        return azimuthal
    elif HOURANGLE > 12: 
        return (360 - azimuthal)
    
def append_to_excel(data, filename="C:\\Users\\risha\\Desktop\\Sunrise\\Data1.xlsx"):
    try:
        workbook = load_workbook(filename)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Time", "Solar Noon" ,"Hour Angle", "Declination Angle", "Elevation Angle", "Azimuthal Angle"])
    sheet.append(data)
    workbook.save(filename)

def simulate(latitude, longitude):
    local_time = time.localtime()
    year = local_time[0]
    month = local_time[1]
    day = local_time[2]
    hour = local_time[3]
    minutes = local_time[4]
    seconds = local_time[5]
    yDay = local_time[7]

    local_time_string = str(hour) + ":" + str(minutes) + ":" + str(seconds)
    julian_day, julian_centuary = julian_day_centuary(year, month, day)
    geo_mean_long, geo_mean_anom = get_geomMeanLongAnom(julian_centuary)
    escenEarthOrbit = get_escentEarthOrbit(julian_centuary)
    suntrueLong, suntrueAnom = get_suntrueLong_suntrueAnom(geo_mean_long, geo_mean_anom, julian_centuary)
    sunRadVector, sunLongApp = get_sunRadVector_sunLongApp(escenEarthOrbit, suntrueAnom, suntrueLong, julian_centuary)
    MOE, MOEcorr = get_mean_obliq_ecliptic_and_correction(julian_centuary)
    sunrtAsccen, sundeclination = get_sun_rt_Asccen_and_sun_decl(sunLongApp, MOEcorr)
    var_y = get_var_y(MOEcorr)
    EOTSN = get_EOTSN(var_y, geo_mean_long, geo_mean_anom, escenEarthOrbit)
    ha_sunrise = HA_sunrise(sundeclination, latitude)
    solarNoon, solar_noon_timeString = solar_noon(EOTSN, longitude)
    tplm = get_time_past_local_midnight(hour, minutes, seconds)
    true_solar_time = get_true_solar_time(tplm, EOTSN, longitude)
    d = local_time.tm_yday + (local_time.tm_hour/24) + (local_time.tm_min/(60*24)) + (local_time.tm_sec/(3600*24)) - 2
    EOT = get_EOT(d)
    local_solar_time = get_local_solar_time(local_time, EOT)
    hour_angle = get_hourAngle(local_solar_time)
    declination_angle = get_declinationAngle(d)
    elevation_angle = get_elevationAngle(declination_angle, latitude, hour_angle)
    azimuthal_angle = get_azimuthal(declination_angle, latitude, hour_angle, elevation_angle) - 1.5
    sunrise_timeString, sunset_timeString, sunrise, sunset = get_sunrise_sunset(solarNoon, ha_sunrise)
    print("Data Logging Started.....")

    data = [local_time_string, solar_noon_timeString, declination_angle, hour_angle, elevation_angle, azimuthal_angle]
    append_to_excel(data)
    print("Time: " + local_time_string + "\n" + "Solar Noon: " + solar_noon_timeString + "\n" + "Declination Angle: " + str(declination_angle) + "\n" + "Hour Anlge: " + str(hour_angle) + "\n" + "Elevation Angle: " + str(elevation_angle) + "\n" + "Azimuthal Angle: " + str(azimuthal_angle) + "\n" + "-------------------------------------------" + "\n")
    return azimuthal_angle, elevation_angle, declination_angle, solar_noon_timeString,local_time_string

# Tkinter GUI
def update_simulation():
    latitude = float(latitude_entry.get())
    longitude = float(longitude_entry.get())
    azimuthal_angle, elevation_angle, declination_angle, solar_noon_timeString,local_time_string = simulate(latitude, longitude)
    azimuthal_angle_label_var.set(f"Azimuthal Angle: {azimuthal_angle}")
    hour_angle_label_var.set(f"Hour Angle: {declination_angle}")
    elevation_angle_label_var.set(f"Elevation Angle: {elevation_angle}")
    solar_noon_label_var.set(f"Solar Noon: {solar_noon_timeString}")
    time_string_label_var.set(f"Time: {local_time_string}")
    root.after(1000, update_simulation)

def start_simulation():
    root.after(1000, update_simulation)

root = Tk()
root.title("Solar Angle Simulator")

latitude_label = Label(root, text="Enter the latitude for the place interested in:")
latitude_label.grid(row=0, column=0, padx=10, pady=10)
latitude_entry = Entry(root)
latitude_entry.grid(row=0, column=1, padx=10, pady=10)

longitude_label = Label(root, text="Enter the longitude for the place interested in:")
longitude_label.grid(row=1, column=0, padx=10, pady=10)
longitude_entry = Entry(root)
longitude_entry.grid(row=1, column=1, padx=10, pady=10)

start_button = Button(root, text="Start Simulation", command=start_simulation)
start_button.grid(row=2, column=0, columnspan=2, pady=20)

azimuthal_angle_label_var = StringVar()
hour_angle_label_var = StringVar()
elevation_angle_label_var = StringVar()
solar_noon_label_var = StringVar()
time_string_label_var = StringVar()

time_string_label = Label(root, textvariable=time_string_label_var)
time_string_label.grid(row=3, column=0, columnspan=2, padx=10, pady=5)

azimuthal_angle_label = Label(root, textvariable=azimuthal_angle_label_var)
azimuthal_angle_label.grid(row=4, column=0, columnspan=2, padx=10, pady=5)

hour_angle_label = Label(root, textvariable=hour_angle_label_var)
hour_angle_label.grid(row=5, column=0, columnspan=2, padx=10, pady=5)

elevation_angle_label = Label(root, textvariable=elevation_angle_label_var)
elevation_angle_label.grid(row=6, column=0, columnspan=2, padx=10, pady=5)

solar_noon_label = Label(root, textvariable=solar_noon_label_var)
solar_noon_label.grid(row=7, column=0, columnspan=2, padx=10, pady=5)

root.mainloop()


    